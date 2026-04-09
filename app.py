"""
╔══════════════════════════════════════════════════════════════╗
║  GTD Telecomunicaciones — Dashboard Ejecutivo PMO            ║
║  Control Semanal · MMOO · Starlink · Fibra Óptica            ║
║  Ejecutar: streamlit run app.py                              ║
╚══════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import date, timedelta
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ModuleNotFoundError:
    OPENPYXL_OK = False

# ══════════════════════════════════════════════════════════════
#  PALETA CORPORATIVA GTD
# ══════════════════════════════════════════════════════════════
GTD = {
    "navy":      "#042C53",
    "navy_m":    "#0C447C",
    "blue":      "#0055A4",
    "blue_m":    "#378ADD",
    "blue_l":    "#B5D4F4",
    "blue_xl":   "#E6F1FB",
    "gray_dk":   "#2C2C2A",
    "gray_m":    "#5F5E5A",
    "gray_s":    "#888780",
    "gray_l":    "#D3D1C7",
    "gray_xl":   "#F1EFE8",
    "gray_wkd":  "#ECEAE3",
    "white":     "#FFFFFF",
    "gantt_bg":  "#D9D9D9",
    "gantt_ok":  "#0055A4",
    "gantt_red": "#C0392B",
    "green":     "#1E8449",
    "green_l":   "#D5F5E3",
    "amber":     "#E67E22",
    "amber_l":   "#FDEBD0",
    "red":       "#C0392B",
    "red_l":     "#FADBD8",
    "hoy":       "#E74C3C",
}

# ══════════════════════════════════════════════════════════════
#  CONFIGURACION STREAMLIT  (debe ser la primera llamada a st)
# ══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="GTD · Dashboard PMO",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════
#  DATOS DE DEMO
# ══════════════════════════════════════════════════════════════
def df_demo() -> pd.DataFrame:
    hoy  = date.today()
    base = hoy - timedelta(days=14)
    def fi(n): return pd.Timestamp(base + timedelta(days=n))

    registros = [
        {"Fase":"Inicio",       "Actividades":"Levantamiento FO Quinta Normal",   "Sede":"Quinta Normal",  "Tecnologias":"Fibra Optica","Fecha Inicio":fi(0), "Dias":5,  "Progreso":1.00,"Responsable":"Fabian Penrroz",  "Estado de Salud":"A tiempo"},
        {"Fase":"Inicio",       "Actividades":"Inspeccion Tecnica Maipu",         "Sede":"Maipu",          "Tecnologias":"MMOO",       "Fecha Inicio":fi(2), "Dias":3,  "Progreso":1.00,"Responsable":"Rodrigo Alarcon", "Estado de Salud":"A tiempo"},
        {"Fase":"Planificacion","Actividades":"Diseno Red MMOO Sector Norte",      "Sede":"Quinta Normal",  "Tecnologias":"MMOO",       "Fecha Inicio":fi(3), "Dias":8,  "Progreso":0.70,"Responsable":"Fabian Penrroz",  "Estado de Salud":"En Riesgo"},
        {"Fase":"Planificacion","Actividades":"Config. Enlace Starlink Penalolen", "Sede":"Penalolen",      "Tecnologias":"Starlink",   "Fecha Inicio":fi(4), "Dias":6,  "Progreso":0.50,"Responsable":"Erika Ibarra",    "Estado de Salud":"En Riesgo"},
        {"Fase":"Ejecucion",    "Actividades":"Montaje Antena Starlink Maipu",     "Sede":"Maipu",          "Tecnologias":"Starlink",   "Fecha Inicio":fi(6), "Dias":9,  "Progreso":0.40,"Responsable":"Rodrigo Alarcon", "Estado de Salud":"A tiempo"},
        {"Fase":"Ejecucion",    "Actividades":"Tendido FO Troncal Sector Sur",     "Sede":"Quinta Normal",  "Tecnologias":"Fibra Optica","Fecha Inicio":fi(5), "Dias":12, "Progreso":0.00,"Responsable":"Fabian Penrroz",  "Estado de Salud":"Retrasado"},
        {"Fase":"Ejecucion",    "Actividades":"Instalacion Router MPLS Central",   "Sede":"Santiago Centro","Tecnologias":"MPLS",       "Fecha Inicio":fi(8), "Dias":7,  "Progreso":0.00,"Responsable":"Erika Ibarra",    "Estado de Salud":"Retrasado"},
        {"Fase":"Ejecucion",    "Actividades":"Configuracion Firewall Fortinet",   "Sede":"Maipu",          "Tecnologias":"Fortinet",   "Fecha Inicio":fi(10),"Dias":5,  "Progreso":0.20,"Responsable":"Rodrigo Alarcon", "Estado de Salud":"A tiempo"},
        {"Fase":"Cierre",       "Actividades":"Certificacion Enlace MMOO Torre A", "Sede":"Penalolen",      "Tecnologias":"MMOO",       "Fecha Inicio":fi(14),"Dias":4,  "Progreso":0.00,"Responsable":"Fabian Penrroz",  "Estado de Salud":"A tiempo"},
        {"Fase":"Cierre",       "Actividades":"Entrega y Cierre Tecnico Proyecto", "Sede":"Santiago Centro","Tecnologias":"MMOO",       "Fecha Inicio":fi(17),"Dias":3,  "Progreso":0.00,"Responsable":"Erika Ibarra",    "Estado de Salud":"A tiempo"},
    ]
    df = pd.DataFrame(registros)
    df.rename(columns={"Tecnologias": "Tecnologías", "Dias": "Días"}, inplace=True)
    df["Fecha Final"]      = df.apply(lambda r: r["Fecha Inicio"] + pd.Timedelta(days=r["Días"]-1), axis=1)
    df["Dias Completados"] = (df["Días"] * df["Progreso"]).round(1)
    return df

def df_vacio() -> pd.DataFrame:
    cols = ["Fase","Actividades","Sede","Tecnologías","Fecha Inicio",
            "Días","Progreso","Responsable","Estado de Salud","Fecha Final","Dias Completados"]
    return pd.DataFrame(columns=cols)

# ══════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════
if "proyectos" not in st.session_state:
    st.session_state.proyectos = {
        "GTD Infraestructura Telecom": {
            "descripcion":    "Proyecto demo · MMOO · Starlink · Fibra Óptica",
            "cliente":        "GTD Telecomunicaciones",
            "fecha_creacion": date.today().strftime("%d/%m/%Y"),
            "df":             df_demo(),
        }
    }

if "proyecto_activo" not in st.session_state:
    st.session_state.proyecto_activo = None

# ══════════════════════════════════════════════════════════════
#  CSS GLOBAL
# ══════════════════════════════════════════════════════════════
st.markdown(f"""
<style>
  html, body, [class*="css"] {{
      font-family: 'Segoe UI', Arial, sans-serif;
  }}
  [data-testid="stSidebar"] {{
      background-color: {GTD['navy']};
  }}
  [data-testid="stSidebar"] * {{
      color: {GTD['white']} !important;
  }}
  [data-testid="stSidebar"] label {{
      color: {GTD['blue_l']} !important;
      font-size: 11px !important;
      font-weight: 700 !important;
      text-transform: uppercase;
      letter-spacing: .06em;
  }}
  .block-container {{ padding-top: 1.2rem !important; }}

  .gtd-header {{
      background: {GTD['navy']};
      padding: 18px 24px 14px;
      border-radius: 10px;
      margin-bottom: 18px;
  }}
  .gtd-header h1 {{
      color: {GTD['white']};
      font-size: 20px;
      font-weight: 700;
      margin: 0;
  }}
  .gtd-header p {{
      color: {GTD['blue_l']};
      font-size: 12px;
      margin: 4px 0 0;
  }}

  .kpi {{
      background: {GTD['white']};
      border-radius: 10px;
      padding: 14px 18px;
      border-left: 5px solid {GTD['blue']};
      box-shadow: 0 1px 6px rgba(4,44,83,.07);
      height: 90px;
  }}
  .kpi.red   {{ border-left-color:{GTD['red']};   background:{GTD['red_l']};   }}
  .kpi.amber {{ border-left-color:{GTD['amber']}; background:{GTD['amber_l']}; }}
  .kpi.green {{ border-left-color:{GTD['green']}; background:{GTD['green_l']}; }}
  .kpi-label {{
      font-size: 10px; font-weight: 700;
      text-transform: uppercase; letter-spacing: .08em;
      color: {GTD['gray_m']}; margin-bottom: 3px;
  }}
  .kpi-value {{
      font-size: 30px; font-weight: 800; color: {GTD['navy']}; line-height: 1.1;
  }}
  .kpi-value.red   {{ color: {GTD['red']};   }}
  .kpi-value.green {{ color: {GTD['green']}; }}
  .kpi-value.amber {{ color: {GTD['amber']}; }}
  .kpi-sub {{ font-size: 10px; color: {GTD['gray_s']}; margin-top: 2px; }}

  .stitle {{
      font-size: 12px; font-weight: 700; color: {GTD['navy']};
      text-transform: uppercase; letter-spacing: .08em;
      border-bottom: 2px solid {GTD['blue']};
      padding-bottom: 5px; margin-bottom: 10px;
  }}

  .proj-card {{
      background: {GTD['white']};
      border: 1px solid {GTD['gray_l']};
      border-left: 5px solid {GTD['blue']};
      border-radius: 10px;
      padding: 16px 20px 14px;
      box-shadow: 0 2px 8px rgba(4,44,83,.07);
      margin-bottom: 4px;
  }}
  .proj-name {{
      font-size: 15px; font-weight: 700; color: {GTD['navy']}; margin-bottom: 3px;
  }}
  .proj-meta {{
      font-size: 11px; color: {GTD['gray_m']}; margin-bottom: 2px;
  }}
  .proj-badge {{
      display: inline-block;
      background: {GTD['blue_xl']};
      color: {GTD['blue']};
      font-size: 10px; font-weight: 700;
      padding: 2px 8px; border-radius: 20px;
      margin-top: 4px; margin-right: 4px;
  }}
  .home-header {{
      background: linear-gradient(135deg, {GTD['navy']} 60%, {GTD['navy_m']} 100%);
      padding: 32px 32px 28px;
      border-radius: 14px;
      margin-bottom: 28px;
      text-align: center;
  }}
  .home-header h1 {{
      color: {GTD['white']};
      font-size: 26px; font-weight: 800; margin: 0 0 6px;
  }}
  .home-header p {{
      color: {GTD['blue_l']};
      font-size: 13px; margin: 0;
  }}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
#  CONSTANTES
# ══════════════════════════════════════════════════════════════
FASES   = ["Inicio", "Planificación", "Ejecución", "Cierre"]
TECNOS  = ["Fibra Óptica", "MMOO", "Starlink", "MPLS", "Fortinet", "Otra"]
ESTADOS = ["A tiempo", "En Riesgo", "Retrasado"]


def guardar_df(nombre, df):
    st.session_state.proyectos[nombre]["df"] = df


# ══════════════════════════════════════════════════════════════
#  EXCEL: EXPORTAR / IMPORTAR
# ══════════════════════════════════════════════════════════════
COLS_EXCEL = [
    "Fase", "Actividades", "Sede", "Tecnologias",
    "Fecha Inicio", "Dias", "Progreso", "Responsable", "Estado de Salud",
]
# Mapeo para columnas con acento (nombre interno -> nombre Excel amigable)
COLS_EXCEL_DISPLAY = [
    "Fase", "Actividades", "Sede", "Tecnologias",
    "Fecha Inicio", "Dias", "Progreso", "Responsable", "Estado de Salud",
]
COLS_INTERNOS = [
    "Fase", "Actividades", "Sede", "Tecnologias",
    "Fecha Inicio", "Dias", "Progreso", "Responsable", "Estado de Salud",
]

def _borde_fino():
    lado = Side(style="thin", color="BDBDBD")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def generar_excel_exportacion(df: pd.DataFrame, nombre_proyecto: str) -> bytes:
    """Genera un .xlsx formateado con los datos del proyecto."""
    if not OPENPYXL_OK:
        st.error("openpyxl no está instalado. Agrega 'openpyxl' a requirements.txt.")
        return b""
    wb = Workbook()
    ws = wb.active
    ws.title = "Actividades"

    encabezados = ["Fase","Actividades","Sede","Tecnologias",
                   "Fecha Inicio","Dias","Progreso","Responsable","Estado de Salud"]
    col_df      = ["Fase","Actividades","Sede","Tecnologias",
                   "Fecha Inicio","Dias","Progreso","Responsable","Estado de Salud"]

    # Titulo
    ws.merge_cells("A1:I1")
    ws["A1"] = f"GTD PMO - {nombre_proyecto}"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", fgColor="042C53")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Exportado: {date.today().strftime('%d/%m/%Y')}  -  {len(df)} actividades"
    ws["A2"].font      = Font(italic=True, size=9, color="5F5E5A", name="Arial")
    ws["A2"].fill      = PatternFill("solid", fgColor="E6F1FB")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    for c_idx, col in enumerate(encabezados, 1):
        cell = ws.cell(row=3, column=c_idx, value=col)
        cell.font      = Font(bold=True, size=10, color="FFFFFF", name="Arial")
        cell.fill      = PatternFill("solid", fgColor="0055A4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _borde_fino()
    ws.row_dimensions[3].height = 22

    color_alt = ["FFFFFF", "EBF4FF"]
    col_salud_colors = {
        "A tiempo": "D5F5E3",
        "En Riesgo": "FDEBD0",
        "Retrasado": "FADBD8",
    }
    # Preparar df a exportar: crear df con columnas internas (Tecnologias, Dias)
    if not df.empty:
        df_exp = df.copy()
        # Manejar nombres alternativos
        for col_alt, col_canon in [("Tecnologias","Tecnologias"), ("Dias","Dias")]:
            if col_canon not in df_exp.columns and col_alt in df_exp.columns:
                df_exp[col_canon] = df_exp[col_alt]
    else:
        df_exp = pd.DataFrame(columns=col_df)
    # Asegurar que todas las columnas existan
    for c in col_df:
        if c not in df_exp.columns:
            df_exp[c] = ""

    for r_idx, (_, row) in enumerate(df_exp.iterrows(), 1):
        excel_row  = r_idx + 3
        fondo_base = color_alt[r_idx % 2]
        salud_val  = str(row.get("Estado de Salud", ""))

        for c_idx, (enc, col_src) in enumerate(zip(encabezados, col_df), 1):
            val = row.get(col_src, "")
            if enc == "Fecha Inicio" and hasattr(val, "strftime"):
                val = val.strftime("%d/%m/%Y")
            elif enc == "Progreso":
                try:
                    val = f"{int(float(val)*100)}%"
                except Exception:
                    val = "0%"
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font      = Font(size=10, name="Arial")
            cell.alignment = Alignment(
                horizontal="center" if enc in ("Fase","Dias","Progreso","Fecha Inicio","Estado de Salud") else "left",
                vertical="center"
            )
            cell.border = _borde_fino()
            if enc == "Estado de Salud" and salud_val in col_salud_colors:
                cell.fill = PatternFill("solid", fgColor=col_salud_colors[salud_val])
            else:
                cell.fill = PatternFill("solid", fgColor=fondo_base)
        ws.row_dimensions[excel_row].height = 18

    anchos = [15, 38, 20, 14, 14, 8, 10, 20, 15]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Hoja plantilla
    wi = wb.create_sheet("Plantilla")
    wi.merge_cells("A1:I1")
    wi["A1"] = "PLANTILLA PARA IMPORTAR ACTIVIDADES"
    wi["A1"].font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    wi["A1"].fill      = PatternFill("solid", fgColor="042C53")
    wi["A1"].alignment = Alignment(horizontal="center", vertical="center")
    wi.row_dimensions[1].height = 26

    instrucciones = [
        ("A3", "INSTRUCCIONES", True),
        ("A4", "1. Completa los datos desde la fila 9 en adelante. NO modificar los encabezados."),
        ("A5", "2. Fecha Inicio: usa formato DD/MM/AAAA  ej: 15/04/2025"),
        ("A6", "3. Progreso: numero entre 0 y 100 (sin simbolo %).  ej: 75"),
        ("A7", "4. Guarda el archivo y cargalo en la app desde la pestana Importar Excel."),
    ]
    for ref, txt, *bold in instrucciones:
        wi.merge_cells(f"{ref}:I{ref[1:]}")
        wi[ref] = txt
        wi[ref].font = Font(bold=bool(bold), size=10, name="Arial",
                            color="042C53" if bold else "2C2C2A")
        wi.row_dimensions[int(ref[1:])].height = 16

    for c_idx, col in enumerate(encabezados, 1):
        cell = wi.cell(row=8, column=c_idx, value=col)
        cell.font      = Font(bold=True, size=10, color="FFFFFF", name="Arial")
        cell.fill      = PatternFill("solid", fgColor="0055A4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _borde_fino()
    wi.row_dimensions[8].height = 22

    ayudas = {
        "Fase":            " | ".join(FASES),
        "Tecnologias":     " | ".join(TECNOS),
        "Estado de Salud": " | ".join(ESTADOS),
        "Progreso":        "Numero 0-100",
        "Dias":            "Numero entero",
        "Fecha Inicio":    "DD/MM/AAAA",
    }
    for c_idx, col in enumerate(encabezados, 1):
        if col in ayudas:
            cell = wi.cell(row=9, column=c_idx, value=ayudas[col])
            cell.font      = Font(italic=True, size=8, color="5F5E5A", name="Arial")
            cell.fill      = PatternFill("solid", fgColor="E6F1FB")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = _borde_fino()
    wi.row_dimensions[9].height = 30

    ejemplo = ["Ejecucion","Ejemplo: Tendido FO Norte","Quinta Normal",
               "Fibra Optica", date.today().strftime("%d/%m/%Y"),
               7, 30, "Juan Perez", "A tiempo"]
    for c_idx, val in enumerate(ejemplo, 1):
        cell = wi.cell(row=10, column=c_idx, value=val)
        cell.font      = Font(italic=True, size=9, color="888780", name="Arial")
        cell.fill      = PatternFill("solid", fgColor="F1EFE8")
        cell.alignment = Alignment(
            horizontal="center" if c_idx in [1,5,6,7,9] else "left")
        cell.border    = _borde_fino()
    wi.row_dimensions[10].height = 18

    for i, w in enumerate(anchos, 1):
        wi.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_desde_excel(archivo) -> tuple:
    """Lee un Excel subido. Retorna (df_valido_o_None, lista_errores)."""
    errores = []
    try:
        df_raw = pd.read_excel(archivo, sheet_name=0, header=None)
    except Exception as e:
        return None, [f"No se pudo leer el archivo: {e}"]

    # Buscar fila de encabezados
    fila_header = None
    for i, row in df_raw.iterrows():
        if "Actividades" in row.values:
            fila_header = i
            break

    if fila_header is None:
        return None, ["No se encontro la fila de encabezados con columna 'Actividades'."]

    df_raw.columns = df_raw.iloc[fila_header]
    df_raw = df_raw.iloc[fila_header + 1:].reset_index(drop=True)

    # Filtrar vacas y filas de ayuda
    df_raw = df_raw[df_raw["Actividades"].notna()].copy()
    df_raw = df_raw[df_raw["Actividades"].astype(str).str.strip() != ""].copy()
    df_raw = df_raw[~df_raw["Actividades"].astype(str).str.startswith("Ejemplo")].copy()

    if df_raw.empty:
        return None, ["El archivo no contiene filas de datos validos."]

    encabezados_req = ["Fase","Actividades","Sede","Tecnologias",
                       "Fecha Inicio","Dias","Progreso","Responsable","Estado de Salud"]
    faltantes = [c for c in encabezados_req if c not in df_raw.columns]
    if faltantes:
        return None, [f"Faltan columnas: {', '.join(faltantes)}"]

    df = df_raw[encabezados_req].copy()

    # Fechas
    df["Fecha Inicio"] = pd.to_datetime(df["Fecha Inicio"], dayfirst=True, errors="coerce")
    n_malas = df["Fecha Inicio"].isna().sum()
    if n_malas > 0:
        errores.append(f"Advertencia: {n_malas} fila(s) con fecha invalida omitidas.")
        df = df[df["Fecha Inicio"].notna()].copy()

    # Progreso
    def norm_prog(val):
        try:
            s = str(val).replace("%","").strip()
            f = float(s)
            return f / 100 if f > 1 else f
        except Exception:
            return 0.0
    df["Progreso"] = df["Progreso"].apply(norm_prog).clip(0, 1)

    # Dias
    df["Dias"] = pd.to_numeric(df["Dias"], errors="coerce").fillna(1).astype(int).clip(1)

    # Calcular Fecha Final
    df["Fecha Final"]      = df.apply(
        lambda r: r["Fecha Inicio"] + pd.Timedelta(days=int(r["Dias"]) - 1), axis=1
    )
    df["Dias Completados"] = (df["Dias"] * df["Progreso"]).round(1)

    # Renombrar columnas sin tilde a con tilde para el sistema interno
    df.rename(columns={"Tecnologias": "Tecnologias", "Dias": "Dias"}, inplace=True)

    # Strings limpios
    for col in ["Fase","Actividades","Sede","Tecnologias","Responsable","Estado de Salud"]:
        df[col] = df[col].astype(str).str.strip()

    # Valores por defecto para columnas controladas
    df.loc[~df["Fase"].isin(FASES),             "Fase"]            = "Ejecucion"
    df.loc[~df["Estado de Salud"].isin(ESTADOS), "Estado de Salud"] = "A tiempo"

    if df.empty:
        return None, errores + ["No quedaron filas validas."]

    return df, errores



# ══════════════════════════════════════════════════════════════
#  PANTALLA 1: GESTION DE PROYECTOS
# ══════════════════════════════════════════════════════════════
def pantalla_proyectos():
    with st.sidebar:
        st.markdown(f"""
        <div style="text-align:center;padding:18px 0 18px;">
            <div style="font-size:30px;">📡</div>
            <div style="font-size:14px;font-weight:700;color:{GTD['white']};">GTD Telecomunicaciones</div>
            <div style="font-size:10px;color:{GTD['blue_l']};">Dashboard Ejecutivo PMO</div>
        </div>
        <hr style="border-color:rgba(255,255,255,.15);margin:0 0 14px;">
        <div style="font-size:10px;color:{GTD['blue_l']};margin-top:8px;">
        Actualizado: {date.today().strftime('%d/%m/%Y')}
        </div>
        """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="home-header">
      <div style="font-size:36px;margin-bottom:8px;">📁</div>
      <h1>Gestión de Proyectos PMO</h1>
      <p>GTD Telecomunicaciones · Selecciona o crea un proyecto para comenzar la planificación</p>
    </div>
    """, unsafe_allow_html=True)

    proyectos = st.session_state.proyectos
    n_proy    = len(proyectos)

    st.markdown(f'<div class="stitle">📂 Proyectos existentes ({n_proy})</div>',
                unsafe_allow_html=True)

    if n_proy == 0:
        st.info("No hay proyectos creados todavía. Crea el primero usando el formulario de abajo.")
    else:
        for nombre, meta in proyectos.items():
            df_p       = meta["df"]
            n_act      = len(df_p)
            avance     = 0.0
            if n_act > 0 and df_p["Días"].sum() > 0:
                avance = round(
                    (df_p["Progreso"] * df_p["Días"]).sum() / df_p["Días"].sum() * 100, 1
                )
            retrasadas  = int((df_p["Estado de Salud"] == "Retrasado").sum()) if n_act > 0 else 0
            color_borde = GTD["red"] if retrasadas > 0 else GTD["blue"]
            badge_ret   = (
                f"<span class='proj-badge' style='background:{GTD['red_l']};color:{GTD['red']};'>"
                f"⚠️ {retrasadas} retrasadas</span>"
                if retrasadas > 0 else ""
            )

            st.markdown(f"""
            <div class="proj-card" style="border-left-color:{color_borde};">
              <div class="proj-name">📋 {nombre}</div>
              <div class="proj-meta">👤 Cliente: {meta['cliente']}
                &nbsp;·&nbsp; 📅 Creado: {meta['fecha_creacion']}</div>
              <div class="proj-meta" style="margin-bottom:6px;">{meta['descripcion']}</div>
              <span class="proj-badge">🗂 {n_act} actividades</span>
              <span class="proj-badge">📈 Avance {avance}%</span>
              {badge_ret}
            </div>
            """, unsafe_allow_html=True)

            btn1, btn2, _ = st.columns([1.2, 1.4, 6])
            with btn1:
                if st.button("🚀 Abrir", key=f"open_{nombre}",
                             use_container_width=True, type="primary"):
                    st.session_state.proyecto_activo = nombre
                    st.rerun()
            with btn2:
                if st.button("🗑️ Eliminar", key=f"del_{nombre}", use_container_width=True):
                    st.session_state[f"confirm_del_{nombre}"] = True
                    st.rerun()

            if st.session_state.get(f"confirm_del_{nombre}", False):
                st.warning(
                    f"¿Eliminar el proyecto **{nombre}** y todas sus actividades? "
                    "Esta acción no se puede deshacer."
                )
                c1, c2, _ = st.columns([1.2, 1.2, 5])
                with c1:
                    if st.button("✅ Sí, eliminar", key=f"yes_{nombre}", type="primary"):
                        del st.session_state.proyectos[nombre]
                        st.session_state.pop(f"confirm_del_{nombre}", None)
                        st.rerun()
                with c2:
                    if st.button("❌ Cancelar", key=f"no_{nombre}"):
                        st.session_state.pop(f"confirm_del_{nombre}", None)
                        st.rerun()

            st.markdown("<div style='margin-bottom:12px;'></div>", unsafe_allow_html=True)

    # ── Formulario nuevo proyecto ─────────────────────────────
    st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
    st.markdown('<div class="stitle">➕ Crear nuevo proyecto</div>', unsafe_allow_html=True)

    with st.form("form_nuevo_proyecto", clear_on_submit=True):
        fc1, fc2 = st.columns(2)
        with fc1:
            np_nombre  = st.text_input("Nombre del proyecto *",
                                       placeholder="Ej: Expansion FO Zona Sur 2025")
            np_cliente = st.text_input("Cliente *",
                                       placeholder="Ej: GTD Telecomunicaciones",
                                       value="GTD Telecomunicaciones")
        with fc2:
            np_desc = st.text_area("Descripcion",
                                   placeholder="Breve descripcion del alcance…",
                                   height=90)
            np_demo = st.checkbox("Cargar actividades de demo", value=False,
                                  help="Carga 10 actividades de ejemplo.")

        submitted = st.form_submit_button("Crear proyecto", type="primary")

        if submitted:
            nombre_lim = np_nombre.strip()
            if not nombre_lim:
                st.error("El nombre del proyecto es obligatorio.")
            elif not np_cliente.strip():
                st.error("El cliente es obligatorio.")
            elif nombre_lim in st.session_state.proyectos:
                st.error(f"Ya existe un proyecto con el nombre '{nombre_lim}'.")
            else:
                st.session_state.proyectos[nombre_lim] = {
                    "descripcion":    np_desc.strip() or "Sin descripcion",
                    "cliente":        np_cliente.strip(),
                    "fecha_creacion": date.today().strftime("%d/%m/%Y"),
                    "df":             df_demo() if np_demo else df_vacio(),
                }
                st.success(f"Proyecto '{nombre_lim}' creado correctamente.")
                st.rerun()

    st.markdown("<div style='margin-top:36px;'></div>", unsafe_allow_html=True)
    st.markdown(f"""
    <div style="text-align:center;padding:12px;background:{GTD['navy']};border-radius:8px;
                font-size:11px;color:{GTD['blue_l']};">
      GTD Telecomunicaciones &nbsp;·&nbsp; Dashboard Ejecutivo PMO
      &nbsp;·&nbsp; MMOO · Starlink · Fibra Optica
      &nbsp;|&nbsp; {date.today().strftime('%d/%m/%Y')}
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  NORMALIZACIÓN DE COLUMNAS (compatibilidad Excel/demo)
# ══════════════════════════════════════════════════════════════
def normalizar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    """Asegura nombres canónicos con tilde y tipos correctos."""
    df = df.copy()
    # Retornar inmediatamente si está vacío para evitar errores en apply()
    if df.empty:
        return df
    df.rename(columns={"Tecnologias": "Tecnologías", "Dias": "Días"}, inplace=True)
    # Eliminar columnas duplicadas que puedan surgir del rename (ej: "Dias" y "Días" coexistentes)
    df = df.loc[:, ~df.columns.duplicated(keep="last")]
    if "Días" in df.columns:
        df["Días"] = pd.to_numeric(df["Días"].squeeze(), errors="coerce").fillna(1).astype(int)
    if "Progreso" in df.columns:
        df["Progreso"] = pd.to_numeric(df["Progreso"].squeeze(), errors="coerce").fillna(0).clip(0, 1)
    if "Fecha Inicio" in df.columns:
        df["Fecha Inicio"] = pd.to_datetime(df["Fecha Inicio"].squeeze(), dayfirst=True, errors="coerce")
    if "Fecha Final" in df.columns:
        df["Fecha Final"] = pd.to_datetime(df["Fecha Final"].squeeze(), dayfirst=True, errors="coerce")
    # Recalcular Fecha Final solo si hay filas válidas
    if ("Fecha Final" not in df.columns or df["Fecha Final"].isna().all())             and "Fecha Inicio" in df.columns and "Días" in df.columns:
        mask = df["Fecha Inicio"].notna() & df["Días"].notna()
        df.loc[mask, "Fecha Final"] = df.loc[mask].apply(
            lambda r: r["Fecha Inicio"] + pd.Timedelta(days=int(r["Días"]) - 1), axis=1
        )
    if "Dias Completados" not in df.columns and "Días" in df.columns and "Progreso" in df.columns:
        df["Dias Completados"] = (df["Días"] * df["Progreso"]).round(1)
    return df


# ══════════════════════════════════════════════════════════════
#  PANTALLA 2: DASHBOARD DEL PROYECTO ACTIVO
# ══════════════════════════════════════════════════════════════
def pantalla_dashboard(nombre_proyecto: str):
    meta      = st.session_state.proyectos[nombre_proyecto]
    df_master = normalizar_columnas(meta["df"])

    # ── Sidebar ───────────────────────────────────────────────
    with st.sidebar:
        st.markdown(f"""
        <div style="text-align:center;padding:10px 0 6px;">
            <div style="font-size:22px;">📡</div>
            <div style="font-size:13px;font-weight:700;color:{GTD['white']};">GTD Telecomunicaciones</div>
            <div style="font-size:10px;color:{GTD['blue_l']};">Dashboard Ejecutivo PMO</div>
        </div>
        <hr style="border-color:rgba(255,255,255,.15);margin:6px 0 10px;">
        <div style="font-size:10px;font-weight:700;color:{GTD['blue_l']};
                    text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px;">
          Proyecto activo
        </div>
        <div style="font-size:12px;font-weight:700;color:{GTD['white']};
                    margin-bottom:10px;line-height:1.3;word-break:break-word;">
          {nombre_proyecto}
        </div>
        """, unsafe_allow_html=True)

        if st.button("← Cambiar proyecto", use_container_width=True, key="btn_volver"):
            st.session_state.proyecto_activo = None
            st.rerun()

        st.markdown("<hr style='border-color:rgba(255,255,255,.15);margin:10px 0 14px;'>",
                    unsafe_allow_html=True)

        def opts(col):
            vals = df_master[col].dropna().unique().tolist() if not df_master.empty else []
            return ["Todos"] + sorted(vals)

        f_fase = st.selectbox("Fase",        opts("Fase"),         key="f_fase")
        f_sede = st.selectbox("Sede",        opts("Sede"),         key="f_sede")
        f_tecn = st.selectbox("Tecnologia",  opts("Tecnologías"),  key="f_tecn")
        f_resp = st.selectbox("Responsable", opts("Responsable"),  key="f_resp")

        st.markdown("<hr style='border-color:rgba(255,255,255,.15);margin:14px 0;'>",
                    unsafe_allow_html=True)
        escala = st.radio("Escala de tiempo", ["Días", "Semanas", "Meses"], key="escala")

        st.markdown(
            f"<div style='font-size:10px;color:{GTD['gray_s']};margin-top:10px;'>"
            f"Actualizado: {date.today().strftime('%d/%m/%Y')}</div>",
            unsafe_allow_html=True,
        )

    # ── Filtrar ───────────────────────────────────────────────
    df = df_master.copy()
    if not df.empty:
        if f_fase != "Todos": df = df[df["Fase"]        == f_fase]
        if f_sede != "Todos": df = df[df["Sede"]        == f_sede]
        if f_tecn != "Todos": df = df[df["Tecnologías"] == f_tecn]
        if f_resp != "Todos": df = df[df["Responsable"] == f_resp]

    # ── Header ────────────────────────────────────────────────
    st.markdown(f"""
    <div class="gtd-header">
      <h1>📊 {nombre_proyecto}</h1>
      <p>{meta['descripcion']} &nbsp;·&nbsp; Cliente: {meta['cliente']}
         &nbsp;|&nbsp; {date.today().strftime('%d/%m/%Y')}</p>
    </div>
    """, unsafe_allow_html=True)

    # ── KPIs ──────────────────────────────────────────────────
    def calcular_kpis(df_k):
        if df_k.empty:
            return dict(avance=0, retrasadas=0, riesgo=0, completadas=0,
                        total=0, hito="Sin actividades", hito_fecha="—")
        total  = len(df_k)
        avance = 0.0
        if df_k["Días"].sum() > 0:
            avance = round(
                (df_k["Progreso"] * df_k["Días"]).sum() / df_k["Días"].sum() * 100, 1
            )
        retrasadas  = int((df_k["Estado de Salud"] == "Retrasado").sum())
        riesgo      = int((df_k["Estado de Salud"] == "En Riesgo").sum())
        completadas = int((df_k["Progreso"] == 1.0).sum())
        pend = df_k[df_k["Progreso"] == 0]
        if not pend.empty:
            idx    = pend["Fecha Final"].idxmin()
            hito   = pend.loc[idx, "Actividades"]
            hfecha = pend.loc[idx, "Fecha Final"].strftime("%d/%m/%Y")
        else:
            hito, hfecha = "Sin pendientes", "—"
        return dict(avance=avance, retrasadas=retrasadas, riesgo=riesgo,
                    completadas=completadas, total=total, hito=hito, hito_fecha=hfecha)

    k = calcular_kpis(df)

    c1, c2, c3, c4 = st.columns(4)
    av_cls = "green" if k["avance"] >= 70 else ("amber" if k["avance"] >= 40 else "red")

    with c1:
        st.markdown(f"""
        <div class="kpi {av_cls}">
          <div class="kpi-label">% Avance total</div>
          <div class="kpi-value {av_cls}">{k['avance']}%</div>
          <div class="kpi-sub">{k['completadas']}/{k['total']} actividades listas</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <div class="kpi {'red' if k['retrasadas'] > 0 else ''}">
          <div class="kpi-label">Alertas criticas</div>
          <div class="kpi-value {'red' if k['retrasadas'] > 0 else ''}">{k['retrasadas']}</div>
          <div class="kpi-sub">Tareas en estado Retrasado</div>
        </div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""
        <div class="kpi {'amber' if k['riesgo'] > 0 else ''}">
          <div class="kpi-label">En riesgo</div>
          <div class="kpi-value {'amber' if k['riesgo'] > 0 else ''}">{k['riesgo']}</div>
          <div class="kpi-sub">Requieren atencion inmediata</div>
        </div>""", unsafe_allow_html=True)
    with c4:
        st.markdown(f"""
        <div class="kpi">
          <div class="kpi-label">Proximo hito</div>
          <div style="font-size:12px;font-weight:700;color:{GTD['navy']};margin:3px 0 2px;
                      line-height:1.3;">{k['hito']}</div>
          <div class="kpi-sub">📅 {k['hito_fecha']}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<div style='margin-top:22px;'></div>", unsafe_allow_html=True)

    # ── Gestion de actividades ────────────────────────────────
    with st.expander("➕ Gestionar Actividades — Agregar / Eliminar / Excel", expanded=False):
        tab_add, tab_del, tab_xl = st.tabs(["Agregar actividad", "Eliminar actividades", "📥 Importar / 📤 Exportar Excel"])

        with tab_add:
            st.markdown(
                f"<div style='font-size:11px;color:{GTD['gray_m']};margin-bottom:12px;'>"
                "Completa los campos y presiona Agregar.</div>",
                unsafe_allow_html=True,
            )
            fa1, fa2 = st.columns(2)
            with fa1:
                n_fase  = st.selectbox("Fase",       FASES,  key="n_fase")
                n_act   = st.text_input("Actividad",          key="n_act",
                                        placeholder="Ej: Tendido FO Sector Oriente")
                n_sede  = st.text_input("Sede",               key="n_sede",
                                        placeholder="Ej: Santiago Centro")
                n_tecn  = st.selectbox("Tecnologia", TECNOS, key="n_tecn")
            with fa2:
                n_fi    = st.date_input("Fecha Inicio", value=date.today(), key="n_fi")
                n_dias  = st.number_input("Duracion (dias)", min_value=1, max_value=365,
                                          value=5, step=1, key="n_dias")
                n_prog  = st.slider("Progreso (%)", 0, 100, 0, step=5, key="n_prog")
                n_resp  = st.text_input("Responsable",        key="n_resp",
                                        placeholder="Ej: Nombre Apellido")
                n_est   = st.selectbox("Estado de Salud", ESTADOS, key="n_est")

            if st.button("Agregar actividad", type="primary", key="btn_agregar"):
                if not n_act.strip():
                    st.error("El nombre de la actividad no puede estar vacio.")
                elif not n_sede.strip():
                    st.error("La sede no puede estar vacia.")
                elif not n_resp.strip():
                    st.error("El responsable no puede estar vacio.")
                else:
                    fi_ts = pd.Timestamp(n_fi)
                    ff_ts = fi_ts + pd.Timedelta(days=int(n_dias) - 1)
                    nueva = pd.DataFrame([{
                        "Fase":             n_fase,
                        "Actividades":      n_act.strip(),
                        "Sede":             n_sede.strip(),
                        "Tecnologías":      n_tecn,
                        "Fecha Inicio":     fi_ts,
                        "Días":             int(n_dias),
                        "Progreso":         n_prog / 100,
                        "Responsable":      n_resp.strip(),
                        "Estado de Salud":  n_est,
                        "Fecha Final":      ff_ts,
                        "Dias Completados": round(n_dias * n_prog / 100, 1),
                    }])
                    guardar_df(nombre_proyecto,
                               pd.concat([df_master, nueva], ignore_index=True))
                    st.success(f"Actividad '{n_act.strip()}' agregada.")
                    st.rerun()

        with tab_del:
            st.markdown(
                f"<div style='font-size:11px;color:{GTD['gray_m']};margin-bottom:12px;'>"
                "Selecciona una o mas actividades y presiona Eliminar.</div>",
                unsafe_allow_html=True,
            )
            opciones   = df_master["Actividades"].tolist() if not df_master.empty else []
            a_eliminar = st.multiselect("Actividades a eliminar", options=opciones,
                                        placeholder="Selecciona actividades...",
                                        key="a_eliminar")
            cb1, cb2 = st.columns([1, 4])
            with cb1:
                if st.button("Eliminar seleccionadas", type="primary",
                             disabled=len(a_eliminar) == 0, key="btn_eliminar"):
                    nuevo_df = df_master[
                        ~df_master["Actividades"].isin(a_eliminar)
                    ].reset_index(drop=True)
                    guardar_df(nombre_proyecto, nuevo_df)
                    st.success(f"{len(a_eliminar)} actividad(es) eliminada(s).")
                    st.rerun()
            with cb2:
                if a_eliminar:
                    st.warning(f"Se eliminaran: {', '.join(a_eliminar)}")

        with tab_xl:
            xe1, xe2 = st.columns(2)

            # ── EXPORTAR ─────────────────────────────────────
            with xe1:
                st.markdown(
                    f"<div style='font-size:12px;font-weight:700;color:{GTD['navy']};margin-bottom:8px;'>"                    "📤 Exportar a Excel</div>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f"<div style='font-size:11px;color:{GTD['gray_m']};margin-bottom:10px;'>"                    "Descarga todas las actividades del proyecto en un .xlsx formateado. "                    "Incluye hoja de plantilla para reimportar.</div>",
                    unsafe_allow_html=True,
                )
                df_actual = st.session_state.proyectos[nombre_proyecto]["df"]
                excel_bytes = generar_excel_exportacion(df_actual, nombre_proyecto)
                st.download_button(
                    label="⬇️ Descargar Excel del proyecto",
                    data=excel_bytes,
                    file_name=f"GTD_PMO_{nombre_proyecto.replace(' ','_')}_{date.today().isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                    key="btn_exportar",
                )
                if not df_actual.empty:
                    st.caption(f"El archivo incluira {len(df_actual)} actividades.")
                else:
                    st.caption("El proyecto no tiene actividades aun. Puedes descargar la plantilla vacia.")

            # ── IMPORTAR ─────────────────────────────────────
            with xe2:
                st.markdown(
                    f"<div style='font-size:12px;font-weight:700;color:{GTD['navy']};margin-bottom:8px;'>"                    "📥 Importar desde Excel</div>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f"<div style='font-size:11px;color:{GTD['gray_m']};margin-bottom:10px;'>"                    "Sube un .xlsx con el formato de la plantilla. Puedes agregar o reemplazar las actividades existentes.</div>",
                    unsafe_allow_html=True,
                )
                archivo_subido = st.file_uploader(
                    "Selecciona archivo .xlsx",
                    type=["xlsx","xls"],
                    key="file_uploader_xl",
                )
                modo_import = st.radio(
                    "Modo de importacion",
                    ["Agregar a las existentes", "Reemplazar todas las actividades"],
                    key="modo_import",
                    horizontal=True,
                )
                if archivo_subido is not None:
                    df_importado, errores_imp = importar_desde_excel(archivo_subido)

                    if errores_imp:
                        for err in errores_imp:
                            if err.startswith("Advertencia"):
                                st.warning(err)
                            else:
                                st.error(err)

                    if df_importado is not None:
                        st.success(f"Archivo valido: {len(df_importado)} actividades listas para importar.")

                        # Preview
                        with st.expander("Vista previa de los datos", expanded=False):
                            preview = df_importado[["Fase","Actividades","Sede","Responsable","Estado de Salud"]].copy()
                            st.dataframe(preview, use_container_width=True, height=200)

                        if st.button("✅ Confirmar importacion", type="primary", key="btn_confirmar_import"):
                            df_actual_imp = st.session_state.proyectos[nombre_proyecto]["df"]
                            # Adaptar nombres de columnas del Excel a nombres internos del sistema
                            df_importado_sys = df_importado.rename(columns={
                                "Tecnologias": "Tecnologias",
                                "Dias": "Dias",
                            })
                            # Columnas internas con acento
                            if "Tecnologias" not in df_importado_sys.columns and "Tecnologias" in df_importado_sys.columns:
                                df_importado_sys = df_importado_sys.rename(columns={"Tecnologias":"Tecnologias"})
                            if "Dias" not in df_importado_sys.columns and "Dias" in df_importado_sys.columns:
                                df_importado_sys = df_importado_sys.rename(columns={"Dias":"Dias"})

                            if modo_import == "Reemplazar todas las actividades":
                                guardar_df(nombre_proyecto, df_importado_sys.reset_index(drop=True))
                                st.success(f"Proyecto actualizado: {len(df_importado_sys)} actividades cargadas.")
                            else:
                                merged = pd.concat([df_actual_imp, df_importado_sys], ignore_index=True)
                                guardar_df(nombre_proyecto, merged)
                                st.success(f"{len(df_importado_sys)} actividades agregadas. Total: {len(merged)}.")
                            st.rerun()

    st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
    # Refrescar por si hubo cambios en el expander
    df_master = st.session_state.proyectos[nombre_proyecto]["df"]

    # ── Carta Gantt ───────────────────────────────────────────
    def construir_gantt(df_g: pd.DataFrame, esc: str) -> go.Figure:
        fig = go.Figure()
        if df_g.empty:
            fig.add_annotation(
                text="Sin actividades para los filtros seleccionados.",
                xref="paper", yref="paper", x=0.5, y=0.5,
                showarrow=False,
                font=dict(size=15, color=GTD["gray_m"], family="Segoe UI"),
            )
            fig.update_layout(
                height=200, paper_bgcolor=GTD["white"], plot_bgcolor=GTD["gray_xl"],
                xaxis=dict(visible=False), yaxis=dict(visible=False),
                margin=dict(l=10, r=10, t=10, b=10),
            )
            return fig

        fecha_min = df_g["Fecha Inicio"].min()
        fecha_max = df_g["Fecha Final"].max()
        hoy_ts    = pd.Timestamp(date.today())

        def gen_ticks(fmin, fmax, esc):
            ticks, labels = [], []
            cur = fmin
            n   = 1
            lim = fmax + pd.Timedelta(days=10)
            while cur <= lim:
                ticks.append(cur)
                if esc == "Días":
                    labels.append(f"D{n}<br>{cur.strftime('%d/%m')}")
                    cur += pd.Timedelta(days=1)
                elif esc == "Semanas":
                    labels.append(f"S{n}<br>{cur.strftime('%d/%m')}")
                    cur += pd.Timedelta(weeks=1)
                else:
                    labels.append(cur.strftime("%b %Y"))
                    mes = cur.month + 1 if cur.month < 12 else 1
                    ano = cur.year if cur.month < 12 else cur.year + 1
                    cur = cur.replace(year=ano, month=mes, day=1)
                n += 1
            return ticks, labels

        ticks, tick_labels = gen_ticks(fecha_min, fecha_max, esc)

        shapes = []
        cur = fecha_min.normalize()
        while cur <= fecha_max + pd.Timedelta(days=1):
            if cur.weekday() == 5:
                shapes.append(dict(
                    type="rect", xref="x", yref="paper",
                    x0=cur, x1=cur + pd.Timedelta(days=2),
                    y0=0, y1=1,
                    fillcolor=GTD["gray_wkd"], opacity=0.55,
                    layer="below", line_width=0,
                ))
            cur += pd.Timedelta(days=1)

        shapes.append(dict(
            type="line", xref="x", yref="paper",
            x0=hoy_ts, x1=hoy_ts, y0=0, y1=1,
            line=dict(color=GTD["hoy"], width=2, dash="dot"),
        ))

        for _, row in df_g.iloc[::-1].reset_index(drop=True).iterrows():
            fi_ts  = row["Fecha Inicio"]
            ff_ts  = row["Fecha Final"]
            prog   = float(row["Progreso"])
            estado = row["Estado de Salud"]
            act    = row["Actividades"]

            if estado == "Retrasado":
                c_base, c_prog = GTD["gantt_red"], "#8B1A1A"
            elif estado == "En Riesgo":
                c_base, c_prog = "#F0A500", "#A06800"
            else:
                c_base, c_prog = GTD["gantt_bg"], GTD["gantt_ok"]

            tooltip = (
                f"<b>{act}</b><br>"
                f"Sede: {row['Sede']} · Fase: {row['Fase']}<br>"
                f"Responsable: {row['Responsable']}<br>"
                f"Inicio: {fi_ts.strftime('%d/%m/%Y')}<br>"
                f"Fin: {ff_ts.strftime('%d/%m/%Y')}<br>"
                f"Duracion: {int(row.get('Días', row.get('Dias', 1)))} dias<br>"
                f"Progreso: {prog*100:.0f}%<br>"
                f"Estado: <b>{estado}</b><br>"
                "<extra></extra>"
            )

            fig.add_trace(go.Bar(
                name="", x=[ff_ts], base=[fi_ts], y=[act], orientation="h",
                marker=dict(color=c_base, line=dict(color="rgba(0,0,0,0.12)", width=0.5)),
                width=0.55, showlegend=False, hovertemplate=tooltip,
            ))

            if prog > 0:
                ff_prog = fi_ts + (ff_ts - fi_ts) * prog
                fig.add_trace(go.Bar(
                    name="", x=[ff_prog], base=[fi_ts], y=[act], orientation="h",
                    marker=dict(color=c_prog, opacity=0.90),
                    width=0.55, showlegend=False,
                    hovertemplate=(
                        f"<b>{act}</b><br>Progreso: {prog*100:.0f}%<br>"
                        f"Dias completados: {int(row.get('Días', row.get('Dias', 1)))*prog:.1f}<br><extra></extra>"
                    ),
                ))

            mid_ts  = fi_ts + (ff_ts - fi_ts) * 0.5
            txt_col = GTD["white"] if (prog > 0.3 or estado in ("Retrasado","En Riesgo")) else GTD["gray_dk"]
            fig.add_annotation(
                x=mid_ts, y=act, text=f"<b>{int(prog*100)}%</b>",
                showarrow=False,
                font=dict(size=10, color=txt_col, family="Segoe UI"),
                xanchor="center", yanchor="middle",
            )

        fig.add_annotation(
            x=hoy_ts, y=1, yref="paper", text="HOY",
            showarrow=False, yanchor="bottom",
            font=dict(size=9, color=GTD["hoy"], family="Segoe UI"), xanchor="center",
        )

        fig.update_layout(
            barmode="overlay",
            height=max(320, len(df_g) * 54 + 100),
            margin=dict(l=10, r=20, t=30, b=50),
            paper_bgcolor=GTD["white"], plot_bgcolor=GTD["white"],
            shapes=shapes, font=dict(family="Segoe UI"),
            xaxis=dict(
                type="date",
                range=[fecha_min - pd.Timedelta(days=1), fecha_max + pd.Timedelta(days=4)],
                tickvals=ticks, ticktext=tick_labels,
                tickfont=dict(size=10, color=GTD["gray_dk"], family="Segoe UI"),
                gridcolor=GTD["gray_l"], gridwidth=0.5,
                showgrid=True, zeroline=False, showline=True,
                linecolor=GTD["gray_l"], fixedrange=False,
            ),
            yaxis=dict(
                tickfont=dict(size=11, color=GTD["gray_dk"], family="Segoe UI"),
                showgrid=False, autorange=True, showline=False, fixedrange=False,
            ),
            hoverlabel=dict(
                bgcolor=GTD["navy"], font_size=12, font_family="Segoe UI",
                font_color=GTD["white"], bordercolor=GTD["blue_m"],
            ),
        )
        return fig

    st.markdown(f"""
    <div class="stitle">
      📅 Carta Gantt — Escala: {escala}
      &nbsp;&nbsp;
      <span style="font-size:11px;font-weight:400;color:{GTD['gray_m']};">
        ({len(df)} actividades mostradas)
      </span>
    </div>
    """, unsafe_allow_html=True)

    leg = st.columns([1, 1, 1, 1, 5])
    for col, (color, texto) in zip(leg, [
        (GTD["gantt_bg"],  "Duracion total"),
        (GTD["gantt_ok"],  "Progreso"),
        (GTD["gantt_red"], "Retrasado"),
        ("#F0A500",        "En Riesgo"),
    ]):
        with col:
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:6px;font-size:11px;"
                f"color:{GTD['gray_dk']};margin-bottom:8px;'>"
                f"<div style='width:16px;height:12px;background:{color};"
                f"border-radius:2px;border:1px solid rgba(0,0,0,0.1);'></div>"
                f"{texto}</div>",
                unsafe_allow_html=True,
            )

    st.plotly_chart(
        construir_gantt(df, escala),
        use_container_width=True,
        config={
            "displayModeBar": True,
            "modeBarButtonsToRemove": ["select2d", "lasso2d", "autoScale2d"],
            "displaylogo": False,
            "toImageButtonOptions": {
                "format": "png",
                "filename": f"GTD_Gantt_{date.today().isoformat()}",
                "scale": 2,
            },
        },
    )

    # ── Tabla detalle ─────────────────────────────────────────
    st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
    st.markdown('<div class="stitle">📋 Detalle de Actividades</div>', unsafe_allow_html=True)

    if not df.empty:
        df_t = df[["Fase","Actividades","Sede","Tecnologías","Fecha Inicio","Fecha Final",
                   "Días","Progreso","Responsable","Estado de Salud"]].copy()
        df_t["Fecha Inicio"] = df_t["Fecha Inicio"].dt.strftime("%d/%m/%Y")
        df_t["Fecha Final"]  = df_t["Fecha Final"].dt.strftime("%d/%m/%Y")
        df_t["Progreso"]     = (df_t["Progreso"] * 100).round(0).astype(int).astype(str) + "%"

        def col_fila(row):
            bg = {"Retrasado": "#FADBD8", "En Riesgo": "#FDEBD0"}.get(row["Estado de Salud"], "")
            return [f"background-color:{bg}" if bg else "" for _ in row]

        styled = (
            df_t.style.apply(col_fila, axis=1)
            .set_properties(**{"font-size": "12px", "font-family": "Segoe UI"})
            .set_table_styles([{
                "selector": "th",
                "props": [
                    ("background-color", GTD["navy"]),
                    ("color", GTD["white"]),
                    ("font-size", "11px"),
                    ("font-weight", "700"),
                    ("text-transform", "uppercase"),
                    ("letter-spacing", "0.05em"),
                ],
            }])
        )
        st.dataframe(styled, use_container_width=True, height=360)
    else:
        st.info("No hay actividades para los filtros seleccionados.")

    # ── Graficos resumen ──────────────────────────────────────
    st.markdown("<div style='margin-top:18px;'></div>", unsafe_allow_html=True)
    st.markdown('<div class="stitle">📊 Resumen Ejecutivo</div>', unsafe_allow_html=True)

    if not df.empty:
        cg1, cg2, cg3 = st.columns(3)

        with cg1:
            cnt = df["Estado de Salud"].value_counts().reset_index()
            cnt.columns = ["Estado", "N"]
            c_mapa = {"A tiempo": GTD["green"], "En Riesgo": GTD["amber"], "Retrasado": GTD["red"]}
            fig1 = go.Figure(go.Pie(
                labels=cnt["Estado"], values=cnt["N"],
                marker_colors=[c_mapa.get(e, GTD["gray_s"]) for e in cnt["Estado"]],
                hole=0.55,
                textfont=dict(family="Segoe UI", size=12),
                hovertemplate="<b>%{label}</b><br>%{value} tareas (%{percent})<extra></extra>",
            ))
            fig1.update_layout(
                title=dict(text="Estado de salud", font=dict(size=13, color=GTD["navy"]), x=0.5),
                margin=dict(t=40, b=10, l=10, r=10), height=240,
                paper_bgcolor=GTD["white"],
                legend=dict(font=dict(size=11, family="Segoe UI")),
            )
            st.plotly_chart(fig1, use_container_width=True, config={"displayModeBar": False})

        with cg2:
            af = df.groupby("Fase").apply(
                lambda g: round((g["Progreso"] * g["Días"]).sum() / g["Días"].sum() * 100, 1)
                if g["Días"].sum() > 0 else 0.0
            ).reset_index()
            af.columns = ["Fase", "Avance"]
            orden = ["Inicio", "Planificación", "Ejecución", "Cierre"]
            af["Fase"] = pd.Categorical(af["Fase"], categories=orden, ordered=True)
            af = af.sort_values("Fase")

            fig2 = go.Figure(go.Bar(
                x=af["Avance"], y=af["Fase"], orientation="h",
                marker=dict(
                    color=[GTD["blue"] if v >= 50 else GTD["amber"] for v in af["Avance"]],
                    line=dict(color="rgba(0,0,0,.08)", width=0.5),
                ),
                text=[f"{v}%" for v in af["Avance"]], textposition="outside",
                textfont=dict(size=11, family="Segoe UI"),
                hovertemplate="<b>%{y}</b><br>Avance: %{x}%<extra></extra>",
            ))
            fig2.update_layout(
                title=dict(text="Avance por fase", font=dict(size=13, color=GTD["navy"]), x=0.5),
                xaxis=dict(range=[0, 118], ticksuffix="%", gridcolor=GTD["gray_l"],
                           tickfont=dict(size=10, family="Segoe UI")),
                yaxis=dict(tickfont=dict(size=11, family="Segoe UI")),
                margin=dict(t=40, b=10, l=10, r=30), height=240,
                paper_bgcolor=GTD["white"], plot_bgcolor=GTD["white"],
            )
            st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar": False})

        with cg3:
            cr = df.groupby("Responsable").agg(
                Tareas=("Actividades", "count"),
                Avance=("Progreso", "mean"),
            ).reset_index()
            cr["Avance"] = (cr["Avance"] * 100).round(1)
            fig3 = go.Figure()
            fig3.add_trace(go.Bar(
                name="Tareas", x=cr["Responsable"], y=cr["Tareas"],
                marker_color=GTD["navy_m"],
                hovertemplate="<b>%{x}</b><br>Tareas: %{y}<extra></extra>",
            ))
            fig3.add_trace(go.Scatter(
                name="Avance %", x=cr["Responsable"], y=cr["Avance"],
                mode="lines+markers",
                marker=dict(color=GTD["gantt_ok"], size=8),
                line=dict(color=GTD["gantt_ok"], width=2),
                yaxis="y2",
                hovertemplate="<b>%{x}</b><br>Avance: %{y}%<extra></extra>",
            ))
            fig3.update_layout(
                title=dict(text="Carga por responsable", font=dict(size=13, color=GTD["navy"]), x=0.5),
                yaxis=dict(title="N Tareas", gridcolor=GTD["gray_l"],
                           tickfont=dict(size=10, family="Segoe UI")),
                yaxis2=dict(title="Avance %", overlaying="y", side="right",
                            range=[0, 115], ticksuffix="%",
                            tickfont=dict(size=10, family="Segoe UI")),
                legend=dict(font=dict(size=10, family="Segoe UI"), orientation="h", y=-0.28),
                xaxis=dict(tickfont=dict(size=10, family="Segoe UI")),
                margin=dict(t=40, b=30, l=10, r=40), height=240,
                paper_bgcolor=GTD["white"], plot_bgcolor=GTD["white"],
            )
            st.plotly_chart(fig3, use_container_width=True, config={"displayModeBar": False})

    # Footer
    st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
    st.markdown(f"""
    <div style="text-align:center;padding:12px;background:{GTD['navy']};border-radius:8px;
                font-size:11px;color:{GTD['blue_l']};">
      GTD Telecomunicaciones &nbsp;·&nbsp; Dashboard Ejecutivo PMO
      &nbsp;·&nbsp; {nombre_proyecto}
      &nbsp;|&nbsp; {date.today().strftime('%d/%m/%Y')}
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  ROUTER PRINCIPAL
# ══════════════════════════════════════════════════════════════
if st.session_state.proyecto_activo is None:
    pantalla_proyectos()
else:
    nombre = st.session_state.proyecto_activo
    if nombre not in st.session_state.proyectos:
        st.session_state.proyecto_activo = None
        st.rerun()
    else:
        pantalla_dashboard(nombre)
